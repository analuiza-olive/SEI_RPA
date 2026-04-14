"""
gerar_oficios.py
----------------
Lê a planilha de busca e agrupa os Procedimentos (col G) por
Delegado (col R) e Delegacia (col S), gerando um arquivo de
texto (.txt) ou Excel resumo para montagem dos ofícios.

Uso:
    python gerar_oficios.py                          # usa o arquivo padrão abaixo
    python gerar_oficios.py minha_planilha.xlsx      # passa o arquivo por argumento
    python gerar_oficios.py planilha.xlsx --formato excel   # saída em Excel
    python gerar_oficios.py planilha.xlsx --formato txt     # saída em TXT (padrão)
"""

import sys
import os
import pandas as pd
from datetime import date

# ──────────────────────────────────────────────
# CONFIGURAÇÕES — ajuste aqui se necessário
# ──────────────────────────────────────────────
ARQUIVO_PADRAO = "busca_DEAIN_DELDIA.xlsx"

# Posição das colunas (base-0); altere se a planilha mudar
COL_PROCEDIMENTO = 6   # coluna G
COL_DELEGADO     = 17  # coluna R
COL_DELEGACIA    = 18  # coluna S


def carregar_dados(caminho: str) -> pd.DataFrame:
    df = pd.read_excel(caminho, dtype=str)
    df.columns = df.columns.str.strip()

    nome_proc  = df.columns[COL_PROCEDIMENTO]
    nome_deleg = df.columns[COL_DELEGADO]
    nome_delec = df.columns[COL_DELEGACIA]

    print(f"  Coluna procedimento : {nome_proc}")
    print(f"  Coluna delegado     : {nome_deleg}")
    print(f"  Coluna delegacia    : {nome_delec}")
    print(f"  Total de linhas     : {len(df)}")

    # Renomeia para nomes fixos internos
    df = df.rename(columns={
        nome_proc:  "procedimento",
        nome_deleg: "delegado",
        nome_delec: "delegacia",
    })

    # Remove linhas sem delegado ou delegacia
    antes = len(df)
    df = df.dropna(subset=["delegado", "delegacia", "procedimento"])
    depois = len(df)
    if antes != depois:
        print(f"  ⚠  {antes - depois} linha(s) ignorada(s) por ter célula vazia em delegado/delegacia/procedimento.")

    # Normaliza texto
    for col in ["delegado", "delegacia", "procedimento"]:
        df[col] = df[col].str.strip()

    return df


def agrupar(df: pd.DataFrame) -> dict:
    """Retorna dict: {(delegado, delegacia): [lista de procedimentos únicos]}"""
    grupos = {}
    for (delegado, delegacia), g in df.groupby(["delegado", "delegacia"], sort=True):
        procs = sorted(g["procedimento"].unique().tolist())
        grupos[(delegado, delegacia)] = procs
    return grupos


def salvar_txt(grupos: dict, caminho_saida: str) -> None:
    hoje = date.today().strftime("%d/%m/%Y")
    with open(caminho_saida, "w", encoding="utf-8") as f:
        f.write(f"AGRUPAMENTO DE PROCEDIMENTOS POR DELEGADO/DELEGACIA\n")
        f.write(f"Gerado em: {hoje}\n")
        f.write("=" * 70 + "\n\n")

        for i, ((delegado, delegacia), procs) in enumerate(grupos.items(), 1):
            f.write(f"OFÍCIO Nº {i:03d}\n")
            f.write(f"Delegado  : {delegado}\n")
            f.write(f"Delegacia : {delegacia}\n")
            f.write(f"Procedimentos ({len(procs)}):\n")
            for p in procs:
                f.write(f"  • {p}\n")
            f.write("\n" + "-" * 70 + "\n\n")

    print(f"  ✔  TXT salvo em: {caminho_saida}")


def salvar_excel(grupos: dict, caminho_saida: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Aba 1: Resumo (uma linha por grupo) ──────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F3864")
    body_font    = Font(name="Arial", size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="BBBBBB")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Nº Ofício", "Delegado", "Delegacia", "Qtd Procedimentos", "Procedimentos"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    ws1.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", fgColor="E8F0FE")

    for i, ((delegado, delegacia), procs) in enumerate(grupos.items(), 1):
        row  = i + 1
        fill = alt_fill if i % 2 == 0 else PatternFill()

        dados = [i, delegado, delegacia, len(procs), " | ".join(procs)]
        for c, val in enumerate(dados, 1):
            cell = ws1.cell(row=row, column=c, value=val)
            cell.font      = body_font
            cell.border    = border
            cell.fill      = fill
            cell.alignment = center if c in (1, 4) else left_wrap

    # Larguras
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 36
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 60
    ws1.freeze_panes = "A2"

    # ── Aba 2: Detalhe (uma linha por procedimento) ───────────────────
    ws2 = wb.create_sheet("Detalhe")

    headers2 = ["Nº Ofício", "Delegado", "Delegacia", "Procedimento"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    row = 2
    for i, ((delegado, delegacia), procs) in enumerate(grupos.items(), 1):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for p in procs:
            for c, val in enumerate([i, delegado, delegacia, p], 1):
                cell = ws2.cell(row=row, column=c, value=val)
                cell.font      = body_font
                cell.border    = border
                cell.fill      = fill
                cell.alignment = left_wrap
            row += 1

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 20
    ws2.freeze_panes = "A2"

    wb.save(caminho_saida)
    print(f"  ✔  Excel salvo em : {caminho_saida}")
    print(f"     • Aba 'Resumo'  — {len(grupos)} grupos (um por ofício)")
    print(f"     • Aba 'Detalhe' — uma linha por procedimento")


def main():
    # Argumentos simples via sys.argv
    args = sys.argv[1:]

    arquivo_entrada = ARQUIVO_PADRAO
    formato = "txt"

    for a in args:
        if a.lower() in ("--formato", "--format"):
            pass  # próximo arg é o valor
        elif a.lower() in ("excel", "xlsx"):
            formato = "excel"
        elif a.lower() in ("txt", "texto"):
            formato = "txt"
        elif a.endswith(".xlsx") or a.endswith(".xls"):
            arquivo_entrada = a

    # Verifica se --formato foi passado com valor
    for i, a in enumerate(args):
        if a.lower() == "--formato" and i + 1 < len(args):
            formato = args[i + 1].lower()

    if not os.path.exists(arquivo_entrada):
        print(f"ERRO: Arquivo '{arquivo_entrada}' não encontrado.")
        print(f"Uso: python {sys.argv[0]} <planilha.xlsx> [--formato txt|excel]")
        sys.exit(1)

    print(f"\n📂 Lendo: {arquivo_entrada}")
    df = carregar_dados(arquivo_entrada)

    print(f"\n🔍 Agrupando...")
    grupos = agrupar(df)
    print(f"  {len(grupos)} grupos encontrados (= ofícios a gerar)")

    # Monta nome do arquivo de saída
    base = os.path.splitext(arquivo_entrada)[0]
    hoje = date.today().strftime("%Y%m%d")

    print(f"\n💾 Salvando resultado...")
    if formato == "excel":
        saida = f"{base}_oficios_{hoje}.xlsx"
        salvar_excel(grupos, saida)
    else:
        saida = f"{base}_oficios_{hoje}.txt"
        salvar_txt(grupos, saida)

    print(f"\n✅ Concluído! {len(grupos)} ofícios gerados.\n")


if __name__ == "__main__":
    main()